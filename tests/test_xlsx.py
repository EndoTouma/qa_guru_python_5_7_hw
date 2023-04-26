import os
from openpyxl import load_workbook

from os_path_scripts import RESOURCES_PATH


def test_xlsx():
    resources = os.path.join(RESOURCES_PATH)
    excel_path = os.path.join(resources, 'file_example_XLSX_50.xlsx')

    workbook = load_workbook(excel_path)
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)
    assert sheet.cell(row=3, column=2).value == 'Mara'
